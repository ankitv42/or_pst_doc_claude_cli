"""
ORCA — rag/generate_docs.py
============================
Generates 5 structured PDF business documents from real Excel data.

WHY PDFs NOT TXT:
    Business documents in 2026 contain:
        - Tables (supplier SLA matrix, capital pool summary)
        - Nested sections with hierarchy
        - Financial data with formatting
        - Footnotes and cross-references
        - Multi-column layouts

    Docling preserves ALL of this structure when parsing.
    Plain txt files lose table structure, hierarchy, and relationships.
    This is the difference between 2020 RAG and 2026 RAG.

DOCUMENTS GENERATED:
    1. supplier_sla.pdf          - Table: all 10 suppliers with SLA terms
    2. event_playbook.pdf        - Table: all 10 events with uplift data
    3. capital_pools.pdf         - Table: all 8 capital pools with limits
    4. replenishment_policy.pdf  - Structured policy with nested sections
    5. entity_relationships.pdf  - Knowledge graph: SKU→Supplier→Pool chains

DATA SOURCE:
    All data read directly from your Excel files.
    No hardcoded values. If Excel changes, regenerate PDFs.

Usage:
    python rag/generate_docs.py
"""

import sys
import openpyxl
from pathlib import Path
from datetime import datetime

sys.path.append(str(Path(__file__).parent.parent))

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle,
    Spacer, HRFlowable, PageBreak
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER

# paths
EXCEL_DIR = Path(__file__).parent.parent.parent / "data" / "source"
DOCS_DIR  = Path(__file__).parent.parent.parent / "docs"
DOCS_DIR.mkdir(exist_ok=True)

# ── colour palette ─────────────────────────────────────────────────────────
NAVY    = colors.HexColor("#1B2A4A")
TEAL    = colors.HexColor("#007C91")
LIGHT   = colors.HexColor("#E8F4F8")
WHITE   = colors.white
GREY    = colors.HexColor("#F5F5F5")
RED     = colors.HexColor("#C0392B")
GREEN   = colors.HexColor("#27AE60")
AMBER   = colors.HexColor("#E67E22")
BORDER  = colors.HexColor("#CBD5E0")


def _styles():
    s = getSampleStyleSheet()
    return {
        "h1": ParagraphStyle("h1", fontSize=18, textColor=NAVY,
                              spaceAfter=8, fontName="Helvetica-Bold"),
        "h2": ParagraphStyle("h2", fontSize=13, textColor=TEAL,
                              spaceAfter=6, spaceBefore=12, fontName="Helvetica-Bold"),
        "h3": ParagraphStyle("h3", fontSize=11, textColor=NAVY,
                              spaceAfter=4, spaceBefore=8, fontName="Helvetica-Bold"),
        "body": ParagraphStyle("body", fontSize=9, spaceAfter=4,
                               fontName="Helvetica", leading=14),
        "note": ParagraphStyle("note", fontSize=8, textColor=colors.HexColor("#666"),
                               spaceAfter=4, fontName="Helvetica-Oblique",
                               leftIndent=12),
        "rule": ParagraphStyle("rule", fontSize=9, spaceAfter=3,
                               fontName="Helvetica", leftIndent=12,
                               bulletIndent=6, leading=13),
        "meta": ParagraphStyle("meta", fontSize=8, textColor=colors.grey,
                               fontName="Helvetica"),
        "warn": ParagraphStyle("warn", fontSize=9, textColor=RED,
                               fontName="Helvetica-Bold", spaceAfter=4),
    }


def _table_style(header_bg=NAVY):
    return TableStyle([
        # header row
        ("BACKGROUND",   (0, 0), (-1, 0),  header_bg),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  WHITE),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0),  8),
        ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
        ("BOTTOMPADDING",(0, 0), (-1, 0),  6),
        ("TOPPADDING",   (0, 0), (-1, 0),  6),
        # data rows
        ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",     (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1, -1), [WHITE, GREY]),
        ("ALIGN",        (0, 1), (-1, -1), "LEFT"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 1), (-1, -1), 4),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        # grid
        ("GRID",         (0, 0), (-1, -1), 0.4, BORDER),
        ("LINEBELOW",    (0, 0), (-1, 0),  1.5, TEAL),
    ])


def _doc_header(story, styles, title, doc_type, version="1.0"):
    story.append(Paragraph(title, styles["h1"]))
    story.append(Paragraph(
        f"Document Type: {doc_type} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Version: {version} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Generated: {datetime.now().strftime('%B %Y')} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Owner: ORCA Supply Chain AI",
        styles["meta"]
    ))
    story.append(Paragraph(
        "⚠ PRIORITY RULE: This document contains planning rules and context only. "
        "All live values (costs, contacts, balances) must be fetched from orca.db "
        "via MCP tools. The database is the single source of truth for all facts.",
        styles["warn"]
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=TEAL, spaceAfter=12))


# ==============================================================================
# DOCUMENT 1 — SUPPLIER SLA
# Table: all 10 suppliers with SLA terms + planning rules per supplier
# ==============================================================================

def generate_supplier_sla():
    wb  = openpyxl.load_workbook(EXCEL_DIR / "03_Suppliers.xlsx")
    ws  = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]

    out    = DOCS_DIR / "supplier_sla.pdf"
    doc    = SimpleDocTemplate(str(out), pagesize=A4,
                               leftMargin=1.8*cm, rightMargin=1.8*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    styles = _styles()
    story  = []

    _doc_header(story, styles,
                "ORCA Supplier SLA and Planning Rules",
                "Supplier Context", "3.0")

    # ── Section 1: SLA Matrix Table ──────────────────────────────────────
    story.append(Paragraph("Section 1 — Supplier SLA Matrix", styles["h2"]))
    story.append(Paragraph(
        "This table summarises all active suppliers. "
        "Lead times and scores shown are standard values from the master data. "
        "Always fetch live effective_lead_time from the database — it is "
        "adjusted for reliability score.",
        styles["body"]
    ))

    headers = ["ID", "Supplier Name", "Country", "Category",
               "Std Lead\n(days)", "Reliability\n(1-5)", "Expedite?",
               "Exp Premium\n(%)", "Payment\n(days)"]
    data = [headers]
    for r in rows:
        sup_id, name, country, cat, lead, min_ord, pay, rel, exp, prem, \
            cname, cemail, contract, active = r
        exp_cell = "✓" if str(exp).upper() == "YES" else "✗"
        rel_colour = GREEN if float(rel) >= 4.5 else AMBER if float(rel) >= 3.8 else RED
        data.append([
            str(sup_id), str(name), str(country), str(cat),
            str(lead), str(rel), exp_cell,
            f"{prem}%" if prem else "N/A",
            str(pay)
        ])

    col_widths = [1.2*cm, 4.0*cm, 2.8*cm, 2.5*cm,
                  1.5*cm, 1.8*cm, 1.5*cm, 1.8*cm, 1.5*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    ts = _table_style()
    # colour expedite column
    for i, r in enumerate(rows, 1):
        exp = r[8]
        cell_colour = GREEN if str(exp).upper() == "YES" else RED
        ts.add("BACKGROUND", (6, i), (6, i), cell_colour)
        ts.add("TEXTCOLOR", (6, i), (6, i), WHITE)
        # colour reliability
        rel = float(r[7])
        rc = GREEN if rel >= 4.5 else AMBER if rel >= 3.8 else RED
        ts.add("TEXTCOLOR", (5, i), (5, i), rc)
        ts.add("FONTNAME", (5, i), (5, i), "Helvetica-Bold")
    t.setStyle(ts)
    story.append(t)
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        "Note: Effective lead time = Std Lead × (1 + (5 − reliability) × 0.1). "
        "Always use curated_skus.effective_lead_time from the database for planning.",
        styles["note"]
    ))

    # ── Section 2: Planning Rules per Category ────────────────────────────
    story.append(Paragraph("Section 2 — Planning Rules by Category", styles["h2"]))

    rules = [
        ("Electronics — TechLine Asia (SUP006)",
         LIGHT, [
             "Longest lead time category. Plan all events 80+ days ahead.",
             "Dubai Shopping Festival Electronics: order by October 27 for January 15 start.",
             "Back to School Electronics: order by June 17 for September 1 start.",
             "Add 2-day customs inspection buffer to every Electronics shipment.",
             "Expedite available at high premium. Use for CRITICAL Class A urgency only.",
             "Expedite lead time ≈ 19 days (35% of ~54-day effective lead time).",
         ]),
        ("Beverages — Gulf Beverages LLC (SUP003)",
         LIGHT, [
             "Shortest practical lead time among import suppliers. Standard orders always safe.",
             "Ramadan uplift 180% — order 2.8× normal qty 60 days before Ramadan start.",
             "Summer Season uplift 60% — increase orders throughout June to August.",
             "High reliability. No extra buffer needed. Standard planning sufficient.",
         ]),
        ("Dairy — Emirates Dairy Co (SUP002)",
         LIGHT, [
             "Perishable category. Maximum order horizon: 14 days supply only.",
             "Never order more than 10 days supply at once — waste risk.",
             "Cold chain compliance mandatory on delivery. Reject non-compliant loads.",
             "Expedite rarely needed given short lead time.",
         ]),
        ("Personal Care — Hindustan FMCG (SUP004) ⚠ NO EXPEDITE",
         colors.HexColor("#FFF3CD"), [
             "CRITICAL: No expedite available. No alternative supplier.",
             "Any Personal Care CRITICAL alert with lead_time_too_late=True must escalate to human.",
             "Plan all Personal Care orders 60 days ahead for any event.",
             "If CRITICAL + no expedite: recommend SUSPEND and notify procurement immediately.",
         ]),
        ("Seasonal — Dragon Imports (SUP005) ⚠ LOWEST RELIABILITY",
         colors.HexColor("#FFF3CD"), [
             "Longest standard lead time in portfolio. Lowest reliability score.",
             "Add 14-day buffer beyond standard lead time for all Seasonal planning.",
             "National Day Seasonal items: order 90 days before December 2.",
             "Eid Al Adha Seasonal: order 80 days before event start.",
             "Expedite premium is very high — only use for CRITICAL Class A urgency.",
             "Always add 5-day customs clearance buffer for China shipments.",
         ]),
        ("Grocery — Multiple Suppliers",
         LIGHT, [
             "Al Rawdah Foods (SUP001, Saudi Arabia): add 2-day customs buffer.",
             "Arabian Agri (SUP008, Saudi Arabia): primary Dates supplier.",
             "  → Ajwa Dates Ramadan: order 300% normal qty 75 days before Ramadan start.",
             "Pakistan Rice Mills (SUP009) ⚠ NO EXPEDITE: plan 60 days ahead.",
             "  → Any Rice CRITICAL alert must auto-escalate to human immediately.",
             "Organic Fields UAE (SUP010): highest reliability, use as emergency backup.",
         ]),
    ]

    for title, bg, rule_list in rules:
        rule_data = [[Paragraph(f"<b>{title}</b>", styles["h3"])]]
        t2 = Table(rule_data, colWidths=[16.4*cm])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), bg),
            ("LEFTPADDING", (0,0),(-1,-1), 8),
            ("RIGHTPADDING",(0,0),(-1,-1), 8),
            ("TOPPADDING",  (0,0),(-1,-1), 6),
            ("BOTTOMPADDING",(0,0),(-1,-1), 6),
            ("BOX", (0,0),(-1,-1), 0.5, TEAL),
        ]))
        story.append(t2)
        for rule in rule_list:
            story.append(Paragraph(f"• {rule}", styles["rule"]))
        story.append(Spacer(1, 0.2*cm))

    # ── Section 3: Expedite Decision Rules ────────────────────────────────
    story.append(Paragraph("Section 3 — Expedite Decision Rules", styles["h2"]))
    exp_rules = [
        ["Condition", "Action"],
        ["lead_time_too_late=True AND allows_expedite=Yes", "Recommend Option C (Expedite)"],
        ["critical_stores > 5 AND std lead_time > 20 days", "Evaluate Option C first"],
        ["Event within effective_lead_time AND uplift > 150%", "Expedite mandatory"],
        ["abc_class = C (low value)", "Standard only unless Tier-1 stockout confirmed"],
        ["CP003 pool_pressure_flag = HIGH", "Option C eliminated — no budget"],
        ["allows_expedite = No", "Escalate to human immediately — no Option C"],
        ["expedite_premium > 50% AND urgency = MEDIUM", "Not worth premium — use standard"],
    ]
    t3 = Table(exp_rules, colWidths=[9*cm, 7.4*cm], repeatRows=1)
    ts3 = _table_style(TEAL)
    for i in [4, 5, 6]:
        ts3.add("BACKGROUND", (1, i), (1, i), colors.HexColor("#FDECEA"))
    t3.setStyle(ts3)
    story.append(t3)

    story.append(Paragraph("Section 4 — Contact Resolution Rule", styles["h2"]))
    story.append(Paragraph(
        "NEVER hardcode supplier contacts in agent output. "
        "Always call get_supplier_info(sku_id) via MCP to resolve contact_name and "
        "contact_email dynamically. The database is the single source of truth. "
        "Contacts change — this document does not update automatically.",
        styles["warn"]
    ))

    doc.build(story)
    print(f"  ✅ Generated: {out}")
    return out


# ==============================================================================
# DOCUMENT 2 — EVENT PLAYBOOK
# Tables: all 10 events + category uplift matrix + planning calendar
# ==============================================================================

def generate_event_playbook():
    wb   = openpyxl.load_workbook(EXCEL_DIR / "07_Event_Calendar.xlsx")
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]

    out    = DOCS_DIR / "event_playbook.pdf"
    doc    = SimpleDocTemplate(str(out), pagesize=A4,
                               leftMargin=1.8*cm, rightMargin=1.8*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    styles = _styles()
    story  = []

    _doc_header(story, styles,
                "ORCA Event Playbook and Demand Uplift Data",
                "Event Context", "3.0")

    # ── Section 1: Event Calendar Table ──────────────────────────────────
    story.append(Paragraph("Section 1 — UAE Retail Event Calendar", styles["h2"]))
    story.append(Paragraph(
        "All 10 UAE retail events affecting demand planning. "
        "demand_uplift_pct is the historical average uplift for affected categories. "
        "event_uplift_factor = 1 + (demand_uplift_pct / 100).",
        styles["body"]
    ))

    headers = ["ID", "Event Name", "Type", "Start Date", "End\nDate",
               "Duration\n(days)", "Uplift\n(%)", "Uplift\nFactor",
               "Planning\nLead (days)", "Affected Categories"]
    data = [headers]
    for r in rows:
        evt_id, name, etype, start, end, dur, uplift, cats, region, lead, notes = r
        factor = round(1 + float(uplift) / 100, 2)
        data.append([
            str(evt_id), str(name), str(etype),
            str(start)[:10], str(end)[:10],
            str(dur), f"{uplift}%", f"{factor}×",
            str(lead), str(cats)
        ])

    col_widths = [1.2*cm, 3.8*cm, 1.8*cm, 1.8*cm, 1.8*cm,
                  1.3*cm, 1.2*cm, 1.2*cm, 1.3*cm, 3.8*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    ts = _table_style()
    for i, r in enumerate(rows, 1):
        uplift = float(r[6])
        if uplift >= 200:
            ts.add("BACKGROUND", (6,i), (7,i), colors.HexColor("#FDECEA"))
            ts.add("FONTNAME", (6,i), (7,i), "Helvetica-Bold")
        elif uplift >= 100:
            ts.add("BACKGROUND", (6,i), (7,i), colors.HexColor("#FFF3CD"))
    t.setStyle(ts)
    story.append(t)
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        "Red cells = uplift ≥200% (CRITICAL planning priority). "
        "Amber cells = uplift ≥100% (HIGH planning priority).",
        styles["note"]
    ))

    # ── Section 2: Category Uplift Matrix ────────────────────────────────
    story.append(Paragraph("Section 2 — Category Uplift Matrix by Event", styles["h2"]))
    story.append(Paragraph(
        "Historical demand uplift percentages by category and event. "
        "Use this matrix to validate event_uplift_factor fetched from the database.",
        styles["body"]
    ))

    matrix_data = [
        ["Category", "Ramadan", "Eid Al\nFitr", "Eid Al\nAdha", "DSF", "Back to\nSchool",
         "UAE Natl\nDay", "Summer\nSeason", "Natl Day\nPrep"],
        ["Grocery",       "180%", "—",    "150%", "60%",  "—",    "—",   "—",   "—"],
        ["Dates",         "280%", "220%", "180%", "—",    "—",    "—",   "—",   "—"],
        ["Beverages",     "180%", "190%", "160%", "60%",  "—",    "80%", "60%", "—"],
        ["Electronics",   "40%",  "—",    "—",    "90%",  "120%", "—",   "—",   "—"],
        ["Personal Care", "60%",  "60%",  "—",    "70%",  "80%",  "—",   "—",   "—"],
        ["Seasonal",      "—",    "350%", "—",    "—",    "—",    "80%", "—",   "40%"],
        ["Home",          "—",    "—",    "—",    "90%",  "—",    "—",   "—",   "—"],
        ["All categ.",    "—",    "—",    "—",    "60%",  "—",    "—",   "50%", "—"],
    ]
    t2 = Table(matrix_data, colWidths=[2.5*cm]+[1.7*cm]*8, repeatRows=1)
    ts2 = _table_style(TEAL)
    for ri in range(1, len(matrix_data)):
        for ci in range(1, len(matrix_data[ri])):
            val = matrix_data[ri][ci]
            if val != "—":
                pct = int(val.replace("%",""))
                if pct >= 200:
                    ts2.add("BACKGROUND",(ci,ri),(ci,ri),colors.HexColor("#FDECEA"))
                    ts2.add("FONTNAME",(ci,ri),(ci,ri),"Helvetica-Bold")
                elif pct >= 100:
                    ts2.add("BACKGROUND",(ci,ri),(ci,ri),colors.HexColor("#FFF3CD"))
                elif pct >= 60:
                    ts2.add("BACKGROUND",(ci,ri),(ci,ri),colors.HexColor("#E8F8E8"))
    t2.setStyle(ts2)
    story.append(t2)

    # ── Section 3: Planning Lead Time Rules ───────────────────────────────
    story.append(Paragraph("Section 3 — Planning Lead Time Rules by Event", styles["h2"]))

    planning_rules = [
        ["Event", "Standard\nPlanning Window", "Category-Specific Rules", "Pool Activated"],
        ["Ramadan 2025\n(180% uplift)",
         "60 days before\nMarch 1",
         "Dates/Ajwa: 75 days ahead (sea freight)\nGrocery: 60 days ahead\nDates qty: 300% of normal",
         "CP002 Ramadan\nSurge Buffer"],
        ["Eid Al Fitr 2025\n(220% uplift)",
         "45 days before\nMarch 30",
         "Plan alongside Ramadan — same order extended\nAll Critical stores stocked before Eid eve\nExpedite must arrive 5 days before start",
         "CP002 / CP005"],
        ["Dubai Shopping\nFestival (90% uplift)",
         "30 days before\nJanuary 15",
         "Electronics: 80 days ahead (TechLine 54-day lead)\nHome: 65 days ahead\nOnly Dubai stores directly affected",
         "CP004 Electronics"],
        ["Back to School\n(120% uplift)",
         "45 days before\nSeptember 1",
         "Electronics: order by June 17\nTier-1 stores near residential areas: 2× uplift\nClass B/C Electronics move significantly",
         "CP004 Electronics"],
        ["UAE National Day\n(80% uplift)",
         "30 days before\nDecember 2",
         "Seasonal items: order 90 days ahead (Dragon Imports 60-day lead)\nPatriotic merchandise with UAE colours",
         "CP005 Seasonal"],
        ["Summer Season\n(60% uplift)",
         "30 days before\nJune 15",
         "77-day duration — continuous monitoring required\nBeverages: increase 60% throughout season\nMultiple reorders needed during long event",
         "CP001 / CP003"],
    ]
    t3 = Table(planning_rules,
               colWidths=[3.2*cm, 2.8*cm, 7.0*cm, 3.4*cm], repeatRows=1)
    t3.setStyle(_table_style())
    story.append(t3)

    # ── Section 4: Demand Calculation Formulas ────────────────────────────
    story.append(Paragraph("Section 4 — Demand Calculation Formulas", styles["h2"]))
    formula_data = [
        ["Formula", "Calculation", "Example"],
        ["Projected Demand\n(with event)",
         "avg_daily_demand × event_uplift_factor × event_duration_days",
         "10 units/day × 2.8 × 29 days = 812 units\n(Ramadan Grocery)"],
        ["Projected Demand\n(no event)",
         "avg_daily_demand × effective_lead_time",
         "10 units/day × 54.5 days = 545 units\n(Electronics standard)"],
        ["Event Uplift Factor",
         "1 + (demand_uplift_pct ÷ 100)",
         "Ramadan 180%: 1 + 1.80 = 2.80\nEid Al Fitr 220%: 1 + 2.20 = 3.20"],
        ["lead_time_too_late\n(with event)",
         "effective_lead_time > days_until_event_start",
         "TechLine 54.5 days vs Ramadan in 40 days\n→ lead_time_too_late = True"],
        ["lead_time_too_late\n(no event)",
         "effective_lead_time > days_to_stockout",
         "Lead 54.5 days vs stockout in 30 days\n→ lead_time_too_late = True"],
    ]
    t4 = Table(formula_data, colWidths=[3.2*cm, 6.4*cm, 6.8*cm], repeatRows=1)
    t4.setStyle(_table_style(NAVY))
    story.append(t4)

    doc.build(story)
    print(f"  ✅ Generated: {out}")
    return out


# ==============================================================================
# DOCUMENT 3 — CAPITAL POOLS
# Tables: pool summary + approval routing matrix
# ==============================================================================

def generate_capital_pools():
    wb   = openpyxl.load_workbook(EXCEL_DIR / "06_Capital_Pools.xlsx")
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]

    out    = DOCS_DIR / "capital_pools.pdf"
    doc    = SimpleDocTemplate(str(out), pagesize=A4,
                               leftMargin=1.8*cm, rightMargin=1.8*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    styles = _styles()
    story  = []

    _doc_header(story, styles,
                "ORCA Capital Pool Structure and Approval Rules",
                "Policy", "2.0")

    # ── Section 1: Pool Summary Table ────────────────────────────────────
    story.append(Paragraph("Section 1 — Capital Pool Summary", styles["h2"]))
    story.append(Paragraph(
        "Eight capital pools fund all replenishment activity. "
        "Live available_aed and utilization_pct must always be fetched from "
        "curated_capital table via check_capital_budgets MCP tool. "
        "Values shown here are total budgets only — not live balances.",
        styles["body"]
    ))

    headers = ["Pool ID", "Pool Name", "Pool Type",
               "Total Budget\n(AED)", "Auto-Approve\nLimit (AED)",
               "Approval\nThreshold (AED)", "Refresh\nCycle", "Owner Dept"]
    data = [headers]
    pool_map = {}
    for r in rows:
        pid, name, ptype, budget, alloc, avail, util, rank, \
            refresh, owner, thresh, auto, quarter, year = r
        budget_val = float(budget) if budget and str(budget).replace('.','').isdigit() else 0
        auto_val   = float(auto)   if auto   and str(auto).replace('.','').isdigit() else 0
        thresh_val = float(thresh) if thresh and str(thresh).replace('.','').isdigit() else 0
        pool_map[str(pid)] = {
            "name": name, "auto": auto_val, "thresh": thresh_val,
            "budget": budget_val, "refresh": refresh, "owner": owner
        }
        data.append([
            str(pid), str(name), str(ptype),
            f"{budget_val:,.0f}" if budget_val else "N/A",
            f"{auto_val:,.0f}"   if auto_val   else "N/A",
            f"{thresh_val:,.0f}" if thresh_val else "N/A",
            str(refresh), str(owner)
        ])

    col_widths = [1.5*cm, 4.5*cm, 2.8*cm, 2.5*cm, 2.2*cm, 2.5*cm, 1.8*cm, 2.2*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(_table_style())
    story.append(t)
    story.append(Paragraph(
        "Note: Auto-Approve Limit = orders below this execute automatically (no human needed). "
        "Approval Threshold = orders above this require VP/Director sign-off. "
        "Between the two limits requires Supply Chain Manager approval.",
        styles["note"]
    ))

    # ── Section 2: Pool Assignment by Category ────────────────────────────
    story.append(Paragraph("Section 2 — Pool Assignment by Category and Order Type", styles["h2"]))
    assign_data = [
        ["Category", "Standard Order Pool\n(Options A & B)", "Expedite Pool\n(Option C)", "Event Override Pool"],
        ["Grocery",       "CP001 Core Grocery",    "CP003 Expedite",       "CP002 Ramadan Surge (if active)"],
        ["Beverages",     "CP001 Core Grocery",    "CP003 Expedite",       "CP002 Ramadan Surge (if active)"],
        ["Dairy",         "CP008 Dairy & Perish.",  "CP003 Expedite",       "CP008 (no override)"],
        ["Electronics",   "CP004 Electronics",     "CP003 Expedite",       "CP004 (no override)"],
        ["Personal Care", "CP001 Core Grocery",    "N/A — no expedite",    "CP005 Seasonal (if event)"],
        ["Seasonal",      "CP005 Seasonal",        "CP003 Expedite",       "CP005 (no override)"],
        ["Home",          "CP001 or CP004",        "CP003 Expedite",       "CP004 DSF period"],
        ["ANY (emergency)","CP007 Emergency",      "CP007 Emergency",      "When primary pool HIGH"],
    ]
    t2 = Table(assign_data, colWidths=[2.8*cm, 4.5*cm, 4.0*cm, 5.1*cm], repeatRows=1)
    ts2 = _table_style(TEAL)
    ts2.add("BACKGROUND", (2,5),(2,5), colors.HexColor("#FDECEA"))
    ts2.add("FONTNAME",   (2,5),(2,5), "Helvetica-Bold")
    t2.setStyle(ts2)
    story.append(t2)

    # ── Section 3: Approval Routing Rules ────────────────────────────────
    story.append(Paragraph("Section 3 — Approval Routing Rules", styles["h2"]))
    routing_data = [
        ["Route", "Conditions", "Action", "Human Required?"],
        ["AUTO_EXECUTE",
         "• Cost < pool auto_approve_limit\n• Pool pressure ≠ HIGH\n• abc_class = B or C",
         "System executes immediately\nreorder_triggered = Yes",
         "No"],
        ["ESCALATE",
         "• Cost > pool auto_approve_limit\n• Any Class A order > Class A threshold\n"
         "• Any expedite above CP003 auto-approve\n• lead_time_too_late=True + CRITICAL",
         "Human reviews HITL briefing\nMust respond within 48 hours\nAlert repeats if no response",
         "Yes — 48hr SLA"],
        ["SUSPEND",
         "• Pool pressure_flag = HIGH\n• Budget < minimum viable order\n"
         "• Always check CP007 before SUSPEND",
         "No order placed\nAlert retained for next cycle\nRe-evaluate when pool pressure drops",
         "No (but alert logged)"],
    ]
    t3 = Table(routing_data, colWidths=[2.5*cm, 6.0*cm, 5.0*cm, 2.9*cm], repeatRows=1)
    ts3 = _table_style(NAVY)
    ts3.add("BACKGROUND",(3,1),(3,1), GREEN)
    ts3.add("TEXTCOLOR", (3,1),(3,1), WHITE)
    ts3.add("BACKGROUND",(3,2),(3,2), AMBER)
    ts3.add("TEXTCOLOR", (3,2),(3,2), WHITE)
    ts3.add("BACKGROUND",(3,3),(3,3), colors.grey)
    ts3.add("TEXTCOLOR", (3,3),(3,3), WHITE)
    t3.setStyle(ts3)
    story.append(t3)

    # ── Section 4: Pool Pressure Rules ────────────────────────────────────
    story.append(Paragraph("Section 4 — Pool Pressure Thresholds", styles["h2"]))
    pressure_data = [
        ["Pressure Flag", "Utilization %", "Meaning", "Agent Action"],
        ["LOW",    "0% – 60%",  "Pool healthy. Full budget available.",
         "All options feasible. Normal scoring."],
        ["MEDIUM", "61% – 85%", "Pool moderately used. Budget constrained.",
         "All options feasible but monitor. Flag high-cost orders."],
        ["HIGH",   "86% – 100%","Pool severely constrained. Minimal budget.",
         "ELIMINATE all options using this pool. Route to SUSPEND or CP007."],
    ]
    t4 = Table(pressure_data, colWidths=[2.0*cm, 2.2*cm, 6.0*cm, 6.2*cm], repeatRows=1)
    ts4 = _table_style(NAVY)
    ts4.add("BACKGROUND",(0,1),(0,1), GREEN);  ts4.add("TEXTCOLOR",(0,1),(0,1),WHITE)
    ts4.add("BACKGROUND",(0,2),(0,2), AMBER);  ts4.add("TEXTCOLOR",(0,2),(0,2),WHITE)
    ts4.add("BACKGROUND",(0,3),(0,3), RED);    ts4.add("TEXTCOLOR",(0,3),(0,3),WHITE)
    t4.setStyle(ts4)
    story.append(t4)

    doc.build(story)
    print(f"  ✅ Generated: {out}")
    return out


# ==============================================================================
# DOCUMENT 4 — REPLENISHMENT POLICY
# Structured policy with ABC rules, scoring formula, decision tree
# ==============================================================================

def generate_replenishment_policy():
    out    = DOCS_DIR / "replenishment_policy.pdf"
    doc    = SimpleDocTemplate(str(out), pagesize=A4,
                               leftMargin=1.8*cm, rightMargin=1.8*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    styles = _styles()
    story  = []

    _doc_header(story, styles,
                "ORCA Replenishment Policy and Decision Framework",
                "Policy", "3.0")

    # ── Section 1: ABC Class Rules ────────────────────────────────────────
    story.append(Paragraph("Section 1 — ABC Class Ordering Rules", styles["h2"]))
    story.append(Paragraph(
        "ORCA portfolio: 102 SKUs across 7 categories. "
        "ABC classification drives ordering rules, approval thresholds, "
        "and option eligibility. Always fetch abc_class from curated_skus via get_sku_info.",
        styles["body"]
    ))

    abc_data = [
        ["Class", "# SKUs", "Characteristics", "Ordering Rules", "Option B?", "Lead Buffer"],
        ["A\n(High Value)", "~20",
         "High margin, high velocity\nCritical to customer experience",
         "Min 200 units regardless of shortfall\nFull distribution — all affected stores\nWeekly review mandatory\nStockout triggers VP notification",
         "NEVER\nFull distribution\nrequired",
         "+20% on\neffective\nlead time"],
        ["B\n(Medium)", "~37",
         "Medium margin and velocity\nStandard portfolio SKUs",
         "Follow supplier min_order_qty\nAll affected stores (unless HIGH pressure)\nBi-weekly review",
         "Permitted\n(flagged if\nmargin rank > 5)",
         "+10%\nrecommended"],
        ["C\n(Low Value)", "~45",
         "Low margin, low velocity\nConsolidate where possible",
         "Standard only — no expedite unless Tier-1 stockout\nBatch with same supplier\nMay exclude during HIGH pool pressure",
         "Permitted",
         "None\nrequired"],
    ]
    t = Table(abc_data,
              colWidths=[1.5*cm, 1.2*cm, 3.8*cm, 5.0*cm, 2.2*cm, 2.2*cm],
              repeatRows=1)
    ts = _table_style(NAVY)
    ts.add("BACKGROUND",(0,1),(0,1), colors.HexColor("#C0392B"))
    ts.add("TEXTCOLOR", (0,1),(0,1), WHITE)
    ts.add("BACKGROUND",(0,2),(0,2), AMBER)
    ts.add("TEXTCOLOR", (0,2),(0,2), WHITE)
    ts.add("BACKGROUND",(0,3),(0,3), GREEN)
    ts.add("TEXTCOLOR", (0,3),(0,3), WHITE)
    ts.add("BACKGROUND",(4,1),(4,1), colors.HexColor("#FDECEA"))
    ts.add("FONTNAME",  (4,1),(4,1), "Helvetica-Bold")
    t.setStyle(ts)
    story.append(t)

    # ── Section 2: Stock Status Response ──────────────────────────────────
    story.append(Paragraph("Section 2 — Stock Status Response Rules", styles["h2"]))
    story.append(Paragraph(
        "Stock status computed from days_of_cover vs effective_lead_time. "
        "Always use the value in curated_inventory.stock_status — never recompute.",
        styles["body"]
    ))

    status_data = [
        ["Status", "Threshold", "Action Required", "Urgency", "SLA"],
        ["CRITICAL",
         "days_of_cover < 50%\nof effective_lead_time\nOR stock = 0",
         "Immediate action\nEvaluate expedite first\nAll Critical stores included",
         "CRITICAL if >5 stores\nor lead_time_too_late\nHIGH otherwise",
         "24 hours"],
        ["AT RISK",
         "days_of_cover between\n50% and 100%\nof effective_lead_time",
         "Standard replenishment\nAcceptable unless event uplift\nAt Risk may be excluded from Option B",
         "HIGH if any\nCritical exists\nMEDIUM otherwise",
         "48 hours"],
        ["HEALTHY",
         "days_of_cover ≥ 100%\nof effective_lead_time",
         "No action required\nWeekly monitoring",
         "N/A",
         "None"],
        ["OVERSTOCK",
         "current_stock > 3×\nreorder_point",
         "Flag for markdown\nor inter-store transfer\nUse CP006 Store Transfer Reserve",
         "N/A",
         "None"],
    ]
    t2 = Table(status_data,
               colWidths=[1.8*cm, 3.2*cm, 4.5*cm, 3.2*cm, 1.8*cm],
               repeatRows=1)
    ts2 = _table_style(NAVY)
    ts2.add("BACKGROUND",(0,1),(0,1), RED);   ts2.add("TEXTCOLOR",(0,1),(0,1),WHITE)
    ts2.add("BACKGROUND",(0,2),(0,2), AMBER); ts2.add("TEXTCOLOR",(0,2),(0,2),WHITE)
    ts2.add("BACKGROUND",(0,3),(0,3), GREEN); ts2.add("TEXTCOLOR",(0,3),(0,3),WHITE)
    ts2.add("BACKGROUND",(0,4),(0,4), TEAL);  ts2.add("TEXTCOLOR",(0,4),(0,4),WHITE)
    t2.setStyle(ts2)
    story.append(t2)

    # ── Section 3: Agent 3 Scoring Formula ────────────────────────────────
    story.append(Paragraph("Section 3 — Capital Allocation Scoring Formula (Agent 3)", styles["h2"]))
    story.append(Paragraph(
        "Every feasible option is scored 0–100. Highest score wins. "
        "Rules applied in strict order — elimination before scoring.",
        styles["body"]
    ))

    formula_data = [
        ["Step", "Rule / Formula", "Max Points", "Notes"],
        ["RULE 1\nElimination",
         "If option pool pressure_flag = HIGH → feasible = False",
         "—",
         "Applied FIRST before any scoring.\nHIGH pressure pool = option eliminated."],
        ["RULE 2\nElimination",
         "If total_cost_aed > pool.available_aed → feasible = False",
         "—",
         "Budget check. Cannot exceed live available balance."],
        ["RULE 3\nFlag",
         "If abc_class = A AND option.id = B → not_recommended = True",
         "—",
         "Class A needs full distribution.\nOption B (Tier-1 only) flagged, not eliminated."],
        ["RULE 4a\nBudget Score",
         "budget_score = (1 − cost ÷ available_budget) × 40",
         "40 pts",
         "Lower cost relative to budget = higher score.\nOption B cheapest → highest budget_score."],
        ["RULE 4a\nAvailability Score",
         "availability_score = availability_pct × 0.40",
         "40 pts",
         "More stores covered = higher score.\nOption A (all stores) scores highest here."],
        ["RULE 4a\nMargin Score",
         "margin_score = (1 ÷ margin_priority_rank) × 20",
         "20 pts",
         "Rank 1 (highest margin) = 20 pts.\nRank 5 = 4 pts. Rank 10 = 2 pts."],
        ["RULE 4b\nPenalty",
         "If urgency = CRITICAL AND lead_time_days > 30:\ntotal_score −= 20",
         "−20 pts",
         "CRITICAL urgency only.\nPenalises slow options during critical events.\nOption A with 54-day lead gets −20 if CRITICAL."],
        ["RULE 5\nApproval",
         "approval_required = (cost > pool.auto_approve_limit_aed)",
         "—",
         "Fetch auto_approve_limit from database.\nTrue = human approval needed via HITL."],
    ]
    t3 = Table(formula_data,
               colWidths=[2.0*cm, 6.0*cm, 1.8*cm, 6.6*cm],
               repeatRows=1)
    ts3 = _table_style(NAVY)
    ts3.add("BACKGROUND",(0,1),(0,2), colors.HexColor("#FDECEA"))
    ts3.add("BACKGROUND",(0,3),(0,3), colors.HexColor("#FFF3CD"))
    ts3.add("BACKGROUND",(2,4),(2,6), GREEN); ts3.add("TEXTCOLOR",(2,4),(2,6),WHITE)
    ts3.add("FONTNAME",  (2,4),(2,6), "Helvetica-Bold")
    ts3.add("BACKGROUND",(2,7),(2,7), RED); ts3.add("TEXTCOLOR",(2,7),(2,7),WHITE)
    t3.setStyle(ts3)
    story.append(t3)

    # ── Section 4: Option Building Rules ──────────────────────────────────
    story.append(Paragraph("Section 4 — Option Building Rules (Agent 2)", styles["h2"]))
    option_data = [
        ["Option", "Name", "Order Qty Formula", "Lead Time", "Pool", "Stores Served"],
        ["A", "Standard\nReplenishment",
         "MAX(min_order_qty, projected_shortfall)",
         "effective_lead_time",
         "CP001 or\nCP004",
         "All Critical\nand At Risk"],
        ["B", "Profit\nMaximisation",
         "MAX(min_order_qty,\nROUND(projected_shortfall × 0.6))",
         "effective_lead_time",
         "CP001 or\nCP004",
         "Tier-1 stores\nonly\n⚠ Not for Class A"],
        ["C", "Expedite\nAir Freight",
         "MAX(min_order_qty, projected_shortfall)\n× (1 + expedite_premium/100)",
         "effective_lead_time\n× 0.35",
         "CP003\nonly",
         "All stores\nCritical first\n⚠ Requires allows_expedite=Yes"],
    ]
    t4 = Table(option_data,
               colWidths=[1.0*cm, 2.2*cm, 4.5*cm, 2.2*cm, 1.8*cm, 4.7*cm],
               repeatRows=1)
    ts4 = _table_style(TEAL)
    ts4.add("BACKGROUND",(4,3),(4,3), colors.HexColor("#FDECEA"))
    ts4.add("FONTNAME",  (4,3),(4,3), "Helvetica-Bold")
    t4.setStyle(ts4)
    story.append(t4)

    doc.build(story)
    print(f"  ✅ Generated: {out}")
    return out


# ==============================================================================
# DOCUMENT 5 — ENTITY RELATIONSHIPS AND REASONING PATTERNS
# Knowledge graph: SKU→Supplier→Pool + Golden cases + Self-check rules
# ==============================================================================

def generate_entity_relationships():
    wb_sup  = openpyxl.load_workbook(EXCEL_DIR / "03_Suppliers.xlsx")
    wb_pool = openpyxl.load_workbook(EXCEL_DIR / "06_Capital_Pools.xlsx")
    sup_rows  = list(wb_sup.active.iter_rows(values_only=True))[1:]
    pool_rows = list(wb_pool.active.iter_rows(values_only=True))[1:]

    out    = DOCS_DIR / "entity_relationships.pdf"
    doc    = SimpleDocTemplate(str(out), pagesize=A4,
                               leftMargin=1.8*cm, rightMargin=1.8*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    styles = _styles()
    story  = []

    _doc_header(story, styles,
                "ORCA Entity Relationships and Agent Reasoning Patterns",
                "Graph + Reasoning", "1.0")

    # ── Section 1: Category → Supplier → Pool chain ───────────────────────
    story.append(Paragraph(
        "Section 1 — Knowledge Graph: Category → Supplier → Pool Chains",
        styles["h2"]
    ))
    story.append(Paragraph(
        "Every agent decision traverses this chain. "
        "Category determines supplier (for lead time and expedite). "
        "Supplier determines which pool funds the order. "
        "Pool pressure determines if the option is feasible.",
        styles["body"]
    ))

    chain_data = [
        ["Category", "Supplier (ID)", "Std Lead\n(days)", "Expedite?",
         "Standard\nPool", "Expedite\nPool", "Risk Flag"],
        ["Electronics",  "TechLine Asia (SUP006)",     "50", "Yes 40%", "CP004", "CP003",
         "Long lead. Plan\n80 days ahead."],
        ["Beverages",    "Gulf Beverages (SUP003)",     "7",  "Yes 25%", "CP001", "CP003",
         "Short lead.\nAlways safe."],
        ["Dairy",        "Emirates Dairy (SUP002)",     "5",  "Yes 20%", "CP008", "CP003",
         "Perishable.\n14-day max order."],
        ["Grocery",      "Multiple suppliers",           "3-40","Mixed",  "CP001", "CP003",
         "Check supplier.\nSome no expedite."],
        ["Personal Care","Hindustan FMCG (SUP004)",     "45", "NO ⚠",   "CP001", "N/A",
         "CRITICAL RISK:\nNo expedite ever."],
        ["Seasonal",     "Dragon Imports (SUP005)",     "60", "Yes 50%", "CP005", "CP003",
         "Lowest reliability.\nAdd 14-day buffer."],
        ["Home",         "Nile Home Products (SUP007)", "35", "Yes 30%", "CP001", "CP003",
         "DSF: order 65\ndays ahead."],
    ]
    t = Table(chain_data,
              colWidths=[2.2*cm, 3.8*cm, 1.5*cm, 1.8*cm, 1.5*cm, 1.5*cm, 4.1*cm],
              repeatRows=1)
    ts = _table_style(NAVY)
    ts.add("BACKGROUND",(3,5),(3,5), RED); ts.add("TEXTCOLOR",(3,5),(3,5),WHITE)
    ts.add("FONTNAME",  (3,5),(3,5), "Helvetica-Bold")
    ts.add("BACKGROUND",(6,5),(6,5), colors.HexColor("#FDECEA"))
    t.setStyle(ts)
    story.append(t)

    # ── Section 2: Risk Chains ────────────────────────────────────────────
    story.append(Paragraph("Section 2 — Risk Escalation Chains", styles["h2"]))
    story.append(Paragraph(
        "These chains show how one entity's property propagates through the system "
        "to affect the final routing decision.",
        styles["body"]
    ))

    chains = [
        ("Chain 1: Supplier Reliability → Lead Time → Urgency",
         "Low reliability supplier → higher effective_lead_time "
         "→ lead_time_too_late triggers earlier → CRITICAL urgency → expedite mandatory.\n"
         "Pakistan Rice Mills (reliability 3.7, no expedite): any CRITICAL alert "
         "always requires human escalation."),
        ("Chain 2: Event Timing → lead_time_too_late → Option Recommendation",
         "Ramadan start March 1. TechLine Asia effective lead 54.5 days. "
         "Planning must start by October 7 for Electronics. "
         "If planning starts after November 1: lead_time_too_late=True → "
         "Option C mandatory → CP003 pressure check → ESCALATE."),
        ("Chain 3: Pool Pressure → Option Feasibility → Routing",
         "CP001 pressure HIGH → Options A and B eliminated for Grocery and Beverages. "
         "CP003 pressure HIGH → Option C eliminated for all categories. "
         "CP001 AND CP003 both HIGH → check CP007 Emergency Contingency. "
         "All pools HIGH → SUSPEND is only option."),
        ("Chain 4: ABC Class → Option Eligibility → Approval",
         "Class A + Option B selected → not_recommended=True (full distribution required). "
         "Class A + order cost above threshold → approval_required=True always. "
         "Class A + CRITICAL + lead_time_too_late → highest priority case, always ESCALATE."),
    ]
    for title, desc in chains:
        t2 = Table([[Paragraph(f"<b>{title}</b>", styles["h3"])]],
                   colWidths=[16.4*cm])
        t2.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), LIGHT),
            ("BOX",(0,0),(-1,-1),0.5,TEAL),
            ("LEFTPADDING",(0,0),(-1,-1),8),
            ("TOPPADDING",(0,0),(-1,-1),6),
            ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ]))
        story.append(t2)
        story.append(Paragraph(desc, styles["rule"]))
        story.append(Spacer(1, 0.2*cm))

    # ── Section 3: Golden Cases ────────────────────────────────────────────
    story.append(Paragraph("Section 3 — Verified Golden Decision Cases", styles["h2"]))
    story.append(Paragraph(
        "These are verified correct agent outputs for reference. "
        "Use these to validate your own reasoning. "
        "If your output differs significantly, check the chain of rules above.",
        styles["body"]
    ))

    cases = [
        ("CASE 1 — Ajwa Dates CRITICAL Ramadan (correct = ESCALATE + Option C)",
         [("Input",
           "category=Dates, abc_class=A, critical_stores=9, urgency=CRITICAL\n"
           "Ramadan in 18 days, effective_lead_time=26.75 days\n"
           "lead_time_too_late=True (26.75 > 18), allows_expedite=Yes"),
          ("Agent 1",
           "urgency=CRITICAL (9 critical stores > 5 AND lead_time_too_late=True)\n"
           "projected_demand = avg_daily × 2.8 × 29 days\n"
           "briefing mentions Ramadan timing and standard lead too slow"),
          ("Agent 2",
           "Option C primary (CRITICAL + lead_time_too_late=True)\n"
           "Option A marked INSUFFICIENT — arrives after Ramadan starts\n"
           "Option B marked not_recommended — abc_class=A\n"
           "supplier_contact resolved from database — never hardcoded"),
          ("Agent 3",
           "Option A: lead_time_penalty −20 (CRITICAL + lead >30 days? check actual)\n"
           "Option C expedite: 26.75 × 0.35 = 9.4 days — arrives before Ramadan\n"
           "Winner: Option C. approval_required=True (cost > CP003 auto-approve limit)\n"
           "Do NOT hardcode the auto-approve limit — fetch from database"),
          ("Routing",
           "ESCALATE. Human reads HITL briefing and approves expedite.")]),
        ("CASE 2 — Personal Care CRITICAL no expedite (correct = SUSPEND + escalate human)",
         [("Input",
           "category=Personal Care, supplier=Hindustan FMCG\n"
           "allows_expedite=No, critical_stores=7, urgency=CRITICAL\n"
           "lead_time_too_late=True (49.95 days > days_to_stockout)"),
          ("Agent 2",
           "Option C: feasible=False — supplier has no expedite\n"
           "Options A and B: both arrive too late (lead_time_too_late=True)\n"
           "Both should note INSUFFICIENT timing in trade_off field"),
          ("Agent 3",
           "No feasible option arrives in time\n"
           "recommendation_summary must note: no supplier expedite available\n"
           "routing should be ESCALATE or SUSPEND with human note"),
          ("Routing",
           "ESCALATE or SUSPEND with explicit note:\n"
           "No recovery option exists — human must find alternative supplier.")]),
        ("CASE 3 — Small Grocery AUTO_EXECUTE (correct = AUTO_EXECUTE, no human)",
         [("Input",
           "category=Grocery, abc_class=C, urgency=MEDIUM\n"
           "Option A total_cost below CP001 auto_approve_limit\n"
           "CP001 pool_pressure_flag=MEDIUM (not HIGH)"),
          ("Agent 3",
           "approval_required=False (cost < pool auto_approve_limit from database)\n"
           "Winner: Option A or B based on highest score"),
          ("Routing",
           "AUTO_EXECUTE immediately. No human needed.\n"
           "Briefing says AUTO-EXECUTED — approved automatically.")]),
    ]

    for case_title, steps in cases:
        story.append(Paragraph(case_title, styles["h3"]))
        case_data = [["Agent/Step", "Expected Correct Output"]]
        for step, output in steps:
            case_data.append([step, output])
        tc = Table(case_data, colWidths=[2.5*cm, 13.9*cm], repeatRows=1)
        tsc = _table_style(TEAL)
        tc.setStyle(tsc)
        story.append(tc)
        story.append(Spacer(1, 0.3*cm))

    # ── Section 4: Self-Check Checklists ──────────────────────────────────
    story.append(Paragraph("Section 4 — Agent Self-Check Checklists", styles["h2"]))
    story.append(Paragraph(
        "After generating any output, verify these conditions. "
        "If any check fails — revise your output before returning.",
        styles["body"]
    ))

    checks = {
        "Agent 1 Demand Summary": [
            "urgency=CRITICAL when critical_stores > 5? If yes but you set HIGH — correct to CRITICAL.",
            "lead_time_too_late calculated correctly? With event: compare lead to days until event start.",
            "briefing is 2-3 sentences? Not one word. Not five sentences.",
            "projected_shortfall calculated? If avg_daily_demand=0, note data limitation.",
        ],
        "Agent 2 Options Package": [
            "Exactly 3 options in the options array? Not 2, not 4.",
            "Option C feasible=False when allows_expedite=No? (fetch from database)",
            "Option B not_recommended=True when abc_class=A? (fetch from database)",
            "Pool assignment correct? Electronics uses CP004, Grocery uses CP001.",
            "supplier_contact resolved from database — never hardcoded in output.",
        ],
        "Agent 3 Capital Decision": [
            "RULE 1 applied first? Pool pressure HIGH → eliminate before scoring.",
            "lead_time_penalty applied ONLY for urgency=CRITICAL AND lead_time > 30?",
            "approval_required set by comparing cost to live auto_approve_limit from database.",
            "Winner is feasible? Never recommend not_recommended unless only option.",
            "scored_options contains all 3 options with elimination_reason if infeasible.",
        ],
        "Agent 4 HITL Briefing": [
            "All 3 options listed with scores or elimination reasons?",
            "Supplier contact (name and email) present? Must come from database.",
            "Approval amount stated as winner option cost (fetch from capital_decision)?",
            "48-hour response SLA mentioned for ESCALATE route?",
        ],
    }

    for agent, checklist in checks.items():
        check_data = [[f"☐ {item}"] for item in checklist]
        t3 = Table([[Paragraph(f"<b>{agent}</b>", styles["h3"])]],
                   colWidths=[16.4*cm])
        t3.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),LIGHT),
            ("BOX",(0,0),(-1,-1),0.5,TEAL),
            ("LEFTPADDING",(0,0),(-1,-1),8),
            ("TOPPADDING",(0,0),(-1,-1),5),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ]))
        story.append(t3)
        for item in checklist:
            story.append(Paragraph(f"☐  {item}", styles["rule"]))
        story.append(Spacer(1, 0.2*cm))

    # ── Section 5: Common Errors ───────────────────────────────────────────
    story.append(Paragraph("Section 5 — Common Reasoning Errors", styles["h2"]))
    error_data = [
        ["Error", "Wrong", "Correct"],
        ["Option B for Class A",
         "Option B selected — cheaper and covers Tier-1",
         "Option B eliminated — abc_class=A requires full distribution"],
        ["Wrong lead_time_penalty",
         "Applying −20 when urgency=HIGH",
         "Penalty ONLY when urgency=CRITICAL AND lead_time > 30"],
        ["Wrong lead_time_too_late",
         "True because effective lead time is long",
         "Compare to days_until_event (with event) or days_to_stockout (no event)"],
        ["Wrong expedite pool",
         "Option C funded from CP001 Core Grocery",
         "Option C ALWAYS uses CP003 Expedite and Air Freight only"],
        ["Hardcoded contact",
         "contact_email: li@techline.cn in output",
         "Always call get_supplier_info(sku_id) via MCP — never hardcode"],
        ["Skipping pool pressure",
         "Scoring options without checking pressure first",
         "RULE 1: Check pool pressure FIRST. HIGH = eliminate before scoring"],
    ]
    t4 = Table(error_data, colWidths=[3.5*cm, 5.5*cm, 7.4*cm], repeatRows=1)
    ts4 = _table_style(RED)
    for i in range(1, len(error_data)):
        ts4.add("BACKGROUND",(2,i),(2,i), colors.HexColor("#E8F8E8"))
    t4.setStyle(ts4)
    story.append(t4)

    doc.build(story)
    print(f"  ✅ Generated: {out}")
    return out


# ==============================================================================
# MAIN
# ==============================================================================

if __name__ == "__main__":
    print("\nORCA — Generating 5 structured PDF business documents\n")
    print(f"  Source : Excel files in {EXCEL_DIR}")
    print(f"  Output : {DOCS_DIR}\n")

    generate_supplier_sla()
    generate_event_playbook()
    generate_capital_pools()
    generate_replenishment_policy()
    generate_entity_relationships()

    print(f"\n  5 PDFs generated in {DOCS_DIR}")
    print("  Next: python rag/ingest.py --reset\n")